import argparse
import json
import os
import re
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ---------- Shared utilities ----------

def count_words(statement: str) -> int:
    """
    Same logic as in validate_and_highlight_excel.py:
    - Unescape \\<, \\>, \\~
    - Replace ; . , with spaces
    - Collapse whitespace
    """
    if not isinstance(statement, str):
        statement = str(statement)
    statement = statement.replace("\\<", "<").replace("\\>", ">").replace("\\~", "~")
    cleaned = re.sub(r"[;.,]+", " ", statement)
    cleaned = re.sub(r"\s+", " ", cleaned)
    words = [w.strip() for w in cleaned.split() if w.strip()]
    return len(words)


def is_unwanted_statement(statement: str) -> bool:
    """
    Same header/boilerplate detector as in remove_header_statements.py.
    """
    if not statement or not isinstance(statement, str):
        return False

    normalized = statement.strip()
    if not normalized:
        return False

    lower = normalized.lower()

    # Pattern 1: "Key Inclusion Criteria:" or "Key Exclusion Criteria:" (with optional colon)
    if re.match(r"^key\s+(inclusion|exclusion)\s+criteria\s*:?\s*$", lower):
        return True

    # Pattern 1b: "Main Inclusion Criteria:" or "Main Exclusion Criteria:" (with optional colon)
    if re.match(r"^main\s+(inclusion|exclusion)\s+criteria\s*:?\s*$", lower):
        return True

    # Pattern 2: "Key Inclusion criteria (Core phase)" or "Key Exclusion criteria (Core phase)"
    if re.match(r"^key\s+(inclusion|exclusion)\s+criteria\s*\([^)]+\)\s*:?\s*$", lower):
        return True

    # Pattern 3: "Inclusion Criteria" or "Exclusion Criteria" (with optional colon)
    if re.match(r"^(inclusion|exclusion)\s+criteria\s*:?\s*$", lower):
        return True

    # Pattern 4: "[Inclusion Criteria]" or "[Exclusion Criteria]"
    if re.match(r"^\[(inclusion|exclusion)\s+criteria\]\s*:?\s*$", lower):
        return True

    # Pattern 5: "INCLUSION CRITERIA" or "EXCLUSION CRITERIA" (all caps)
    if normalized.isupper() and re.match(r"^(INCLUSION|EXCLUSION)\s+CRITERIA\s*:?\s*$", normalized):
        return True

    # Pattern 6: "Additional Inclusion Criteria for..." or "Additional Exclusion Criteria for..."
    if re.match(r"^additional\s+(inclusion|exclusion)\s+criteria\s+for\s+.*\s*:?\s*$", lower):
        return True

    # Pattern 7: "Index Case Investigation Inclusion Criteria:" or "Index Case Investigation Exclusion Criteria:"
    if re.match(
        r"^index\s+case\s+investigation\s+(inclusion|exclusion)\s+criteria\s*:?\s*$", lower
    ):
        return True

    return False


def remove_unwanted_statements(
    data: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    cleaned: List[Dict[str, Any]] = []
    removed: List[Dict[str, Any]] = []
    for record in data:
        statement = record.get("statement", "")
        if is_unwanted_statement(statement):
            removed.append(record)
        else:
            cleaned.append(record)
    return cleaned, removed


# ---------- Core pipeline steps ----------

def load_json_records(input_json: str) -> List[Dict[str, Any]]:
    with open(input_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of records")
    return data


def filter_by_nct_limit(
    records: List[Dict[str, Any]], limit_ncts: int | None
) -> List[Dict[str, Any]]:
    if not limit_ncts:
        return records

    seen: set[str] = set()
    allowed_ncts: set[str] = set()

    for rec in records:
        nct = str(rec.get("nctId") or rec.get("nct_id") or rec.get("NCT_ID") or "").strip()
        if not nct or not nct.startswith("NCT"):
            continue
        if nct not in seen:
            seen.add(nct)
            allowed_ncts.add(nct)
            if len(allowed_ncts) >= limit_ncts:
                break

    if not allowed_ncts:
        return []

    filtered = [
        rec
        for rec in records
        if str(rec.get("nctId") or rec.get("nct_id") or rec.get("NCT_ID") or "").strip()
        in allowed_ncts
    ]
    return filtered


def add_word_counts(records: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Return a DataFrame with an added 'word_count' column.
    Assumes each record has at least 'nctId' (or variant) and 'statement'.
    """
    # Normalize NCT column name for convenience
    norm_records: List[Dict[str, Any]] = []
    for rec in records:
        rec = dict(rec)  # shallow copy
        nct = rec.get("nctId") or rec.get("nct_id") or rec.get("NCT_ID") or ""
        rec["nctId"] = str(nct).strip()
        norm_records.append(rec)

    df = pd.DataFrame(norm_records)
    if "statement" not in df.columns:
        raise ValueError("Input JSON records must contain a 'statement' field")

    df["word_count"] = df["statement"].fillna("").astype(str).apply(count_words)
    return df


def save_validated_excel_and_report(
    df: pd.DataFrame, base_out: str, word_threshold: int
) -> None:
    """
    Mirror validate_and_highlight_excel.py:
    - Save CSV with word counts
    - Create Excel file
    - Highlight rows where word_count < threshold
    - Save discrepancy CSV
    """
    csv_path = f"{base_out}_with_word_counts.csv"
    excel_path = f"{base_out}_validated.xlsx"
    disc_path = f"{base_out}_word_count_discrepancies.csv"

    df.to_csv(csv_path, index=False)

    # Excel write
    df.to_excel(excel_path, index=False, engine="openpyxl")
    wb = load_workbook(excel_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    incorrect_rows = []
    discrepancies: List[Dict[str, Any]] = []

    for idx, row in df.iterrows():
        actual_count = int(row["word_count"])
        if actual_count < word_threshold:
            # Excel row index: header is row 1, so data starts at 2
            excel_row = idx + 2
            incorrect_rows.append(excel_row)
            stored_count = row.get("word_count", "N/A")
            discrepancies.append(
                {
                    "row": excel_row,
                    "nctId": row.get("nctId", ""),
                    "type": row.get("type", "N/A"),
                    "stored_word_count": stored_count,
                    "actual_word_count": actual_count,
                    "difference": stored_count - actual_count
                    if isinstance(stored_count, (int, float))
                    else "N/A",
                    "statement_preview": (
                        row.get("statement", "")[:100]
                        + "..."
                        if len(str(row.get("statement", ""))) > 100
                        else row.get("statement", "")
                    ),
                }
            )

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.fill = red_fill

    wb.save(excel_path)

    if discrepancies:
        disc_df = pd.DataFrame(discrepancies)
        disc_df.to_csv(disc_path, index=False)


def write_statement_docs(
    records: List[Dict[str, Any]], docs_dir: str, include_metadata: bool = True
) -> None:
    """
    For each record, write one .txt file under docs_dir/<NCT_ID>/statement_XXX.txt.
    """
    os.makedirs(docs_dir, exist_ok=True)

    # Group by NCT ID
    per_nct: Dict[str, List[Dict[str, Any]]] = {}
    for rec in records:
        nct = str(rec.get("nctId") or rec.get("nct_id") or rec.get("NCT_ID") or "").strip()
        if not nct:
            continue
        per_nct.setdefault(nct, []).append(rec)

    total_ncts = len(per_nct)
    for nct_idx, (nct, recs) in enumerate(per_nct.items(), start=1):
        print(f"[docs] {nct_idx}/{total_ncts} writing statements for {nct}")
        nct_dir = os.path.join(docs_dir, nct)
        os.makedirs(nct_dir, exist_ok=True)
        for i, rec in enumerate(recs, start=1):
            fname = os.path.join(nct_dir, f"statement_{i:03d}.txt")
            statement = str(rec.get("statement", ""))
            lines: List[str] = []
            if include_metadata:
                lines.append(f"NCT_ID: {nct}")
                if "type" in rec:
                    lines.append(f"TYPE: {rec.get('type')}")
                if "word_count" in rec:
                    lines.append(f"WORD_COUNT: {rec.get('word_count')}")
                lines.append("")  # blank line before statement
            lines.append(statement)
            with open(fname, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))


def run_pipeline(
    input_json: str,
    out_prefix: str,
    docs_dir: str,
    limit_ncts: int | None,
    word_threshold: int,
    remove_headers: bool,
) -> None:
    print(f"Loading JSON from: {input_json}")
    records = load_json_records(input_json)
    print(f"Total records in input JSON: {len(records)}")

    all_ncts = sorted(
        {
            str(r.get("nctId") or r.get("nct_id") or r.get("NCT_ID") or "").strip()
            for r in records
            if str(r.get("nctId") or r.get("nct_id") or r.get("NCT_ID") or "").strip()
        }
    )
    print(f"Distinct NCT IDs in JSON: {len(all_ncts)}")
    if all_ncts:
        sample_preview = ", ".join(all_ncts[:5])
        if len(all_ncts) > 5:
            sample_preview += ", ..."
        print(f"Example NCT IDs: {sample_preview}")

    # Optionally limit to first N distinct NCT IDs (for quick testing)
    records = filter_by_nct_limit(records, limit_ncts)
    if not records:
        print("No records after applying NCT limit/filter.")
        return

    distinct_ncts = sorted(
        {
            str(r.get("nctId") or r.get("nct_id") or r.get("NCT_ID") or "").strip()
            for r in records
            if str(r.get("nctId") or r.get("nct_id") or r.get("NCT_ID") or "").strip()
        }
    )
    print(f"Records after NCT limit: {len(records)} (distinct NCTs: {len(distinct_ncts)})")

    # 1) Add word counts & validate
    df = add_word_counts(records)
    print("Added word counts; saving validation CSV/Excel/discrepancy report...")
    save_validated_excel_and_report(df, out_prefix, word_threshold)

    # 2) Optionally remove header / boilerplate statements
    if remove_headers:
        print("Removing header/boilerplate eligibility statements...")
        full_cleaned_records, removed_rows = remove_unwanted_statements(
            df.to_dict(orient="records")
        )
        print(f"Removed {len(removed_rows)} header/boilerplate records.")

        # Prepare minimal cleaned JSON with only the requested fields
        cleaned_for_json: List[Dict[str, Any]] = []
        for rec in full_cleaned_records:
            nct = str(
                rec.get("nctId") or rec.get("nct_id") or rec.get("NCT_ID") or ""
            ).strip()
            statement = str(rec.get("statement", ""))
            type_val = str(rec.get("type", ""))
            cleaned_for_json.append(
                {
                    "nctId": nct,
                    "statement": statement,
                    "type": type_val,
                    "tags": {},
                }
            )

        # Save cleaned JSON (minimal) and removed CSV (full rows for traceability)
        cleaned_json_path = f"{out_prefix}_cleaned.json"
        removed_csv_path = f"{out_prefix}_removed_headers.csv"
        with open(cleaned_json_path, "w", encoding="utf-8") as f:
            json.dump(cleaned_for_json, f, indent=2, ensure_ascii=False)
        if removed_rows:
            pd.DataFrame(removed_rows).to_csv(removed_csv_path, index=False, encoding="utf-8")

        # For docs, keep the richer records (with word_count, etc.) if present
        records_for_docs = full_cleaned_records
    else:
        records_for_docs = df.to_dict(orient="records")

    # 3) Write one .txt doc per statement per NCT ID
    print(f"Writing per-statement docs under: {docs_dir}")
    write_statement_docs(records_for_docs, docs_dir, include_metadata=True)
    print("Pipeline complete.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "End-to-end pipeline: take a single eligibility JSON and "
            "1) add word counts & validate, "
            "2) optionally remove header statements, "
            "3) write one text file per statement per NCT ID."
        )
    )
    parser.add_argument(
        "--input-json",
        "-i",
        required=True,
        help="Path to input JSON (list of records with at least 'nctId' and 'statement').",
    )
    parser.add_argument(
        "--out-prefix",
        "-o",
        default="eligibility_from_json",
        help="Base path/prefix for summary CSV/Excel outputs (default: eligibility_from_json).",
    )
    parser.add_argument(
        "--docs-dir",
        default="statement_docs",
        help="Directory where per-statement text files will be written (default: statement_docs).",
    )
    parser.add_argument(
        "--limit-ncts",
        type=int,
        default=None,
        help=(
            "If set, only process the first N distinct NCT IDs found in the JSON "
            "(e.g. 10 for quick testing)."
        ),
    )
    parser.add_argument(
        "--word-threshold",
        type=int,
        default=30,
        help="Minimum word count threshold used for validation/highlighting (default: 30).",
    )
    parser.add_argument(
        "--remove-headers",
        action="store_true",
        help="If set, remove header/boilerplate eligibility statements before writing docs.",
    )

    args = parser.parse_args()
    run_pipeline(
        input_json=args.input_json,
        out_prefix=args.out_prefix,
        docs_dir=args.docs_dir,
        limit_ncts=args.limit_ncts,
        word_threshold=args.word_threshold,
        remove_headers=args.remove_headers,
    )


if __name__ == "__main__":
    main()


